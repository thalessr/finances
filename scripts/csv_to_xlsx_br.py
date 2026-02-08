#!/usr/bin/env python3
"""
Converte um CSV para XLSX com cabeçalhos em Português do Brasil
e aplica formatação numérica (padrões brasileiros) para valores e moeda.

Uso: python3 scripts/csv_to_xlsx_br.py input.csv output.xlsx
"""
import sys
from pathlib import Path

def main():
    if len(sys.argv) != 3:
        print("Uso: python3 scripts/csv_to_xlsx_br.py input.csv output.xlsx", file=sys.stderr)
        sys.exit(1)

    input_csv = Path(sys.argv[1])
    output_xlsx = Path(sys.argv[2])

    try:
        import pandas as pd
        from openpyxl import load_workbook
    except Exception as e:
        print("Dependências não encontradas (pandas, openpyxl):", e, file=sys.stderr)
        sys.exit(2)

    df = pd.read_csv(input_csv)

    # Mapeamento de cabeçalhos para português brasileiro
    rename_map = {
        "Year": "Ano",
        "Units": "Unidades",
        "Price": "Preço (R$)",
        "Revenue": "Receita (R$)",
        "VariableCost": "Custo Variável (R$)",
        "FixedCost": "Custos Fixos (R$)",
        "Depreciation": "Depreciação (R$)",
        "EBIT": "EBIT (R$)",
        "Taxes": "Imposto (34%) (R$)",
        "NetIncome": "Lucro Líquido (R$)",
        "OperatingCashFlow": "Fluxo Operacional (R$)",
        "CAPEX": "CAPEX (R$)",
        "FreeCashFlow": "Fluxo de Caixa Livre (R$)",
    }

    df = df.rename(columns=rename_map)

    # Escreve para xlsx com pandas (usa openpyxl como engine)
    df.to_excel(output_xlsx, index=False)

    # Aplica formatação das células com openpyxl
    wb = load_workbook(output_xlsx)
    ws = wb.active

    # Localiza colunas por cabeçalho
    header_row = 1
    headers = {cell.value: cell.column for cell in ws[header_row]}

    # Formatos: unidades inteiras, moeda com símbolo R$ e 2 casas decimais
    int_fmt = '#,##0'
    currency_fmt = '"R$"#,##0.00'

    # Aplica formatação por coluna se a coluna existir
    col_formats = {
        "Unidades": int_fmt,
        "Preço (R$)": currency_fmt,
        "Receita (R$)": currency_fmt,
        "Custo Variável (R$)": currency_fmt,
        "Custos Fixos (R$)": currency_fmt,
        "Depreciação (R$)": currency_fmt,
        "EBIT (R$)": currency_fmt,
        "Imposto (34%) (R$)": currency_fmt,
        "Lucro Líquido (R$)": currency_fmt,
        "Fluxo Operacional (R$)": currency_fmt,
        "CAPEX (R$)": currency_fmt,
        "Fluxo de Caixa Livre (R$)": currency_fmt,
    }

    max_row = ws.max_row
    for header, fmt in col_formats.items():
        col = headers.get(header)
        if not col:
            continue
        for row in range(header_row + 1, max_row + 1):
            cell = ws.cell(row=row, column=col)
            # Se valor for numérico, aplica o formato; caso contrário, tenta converter
            try:
                if isinstance(cell.value, str):
                    # tenta converter strings numéricas com ponto decimal
                    v = cell.value.replace(",", ".")
                    cell.value = float(v)
                # aplica formato
                cell.number_format = fmt
            except Exception:
                # ignora células não numéricas
                pass

    wb.save(output_xlsx)
    print(f"Escrito: {output_xlsx}")

if __name__ == "__main__":
    main()

